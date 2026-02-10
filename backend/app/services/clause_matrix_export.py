"""
Clause-matrix Excel exporter.

Converts a :class:`PageIndexDocument` (document tree with extracted
clauses) into an ``.xlsx`` workbook suitable for review and audit.
"""

from io import BytesIO
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.domain.schema import ClauseItem, PageIndexDocument, PageIndexNode


class ClauseMatrixExporter:
    """Export a clause tree to a formatted Excel workbook."""

    # NOTE: Header labels are in Chinese because the generated Excel
    # file is user-facing output.  Changing the language here would
    # alter the deliverable format.
    HEADERS = [
        "层级关系",     # Level / hierarchy
        "章节标题",     # Section title
        "章节概述",     # Section summary
        "条款ID",       # Clause ID
        "条款内容",     # Clause content
        "页码",         # Page number
        "条款类型",     # Clause type
        "结构化信息",   # Structured info
    ]

    # Column widths (in character units)
    COLUMN_WIDTHS: Dict[str, int] = {
        "A": 15,   # hierarchy
        "B": 25,   # section title
        "C": 40,   # section summary
        "D": 20,   # clause ID
        "E": 50,   # clause content
        "F": 10,   # page number
        "G": 20,   # clause type
        "H": 40,   # structured info
    }

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    @staticmethod
    def export_to_excel(document: PageIndexDocument) -> BytesIO:
        """
        Export the document tree to an Excel workbook.

        Args:
            document: Document with clause tree.

        Returns:
            In-memory ``BytesIO`` containing the ``.xlsx`` bytes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "条款矩阵"  # "Clause Matrix" – user-facing sheet name

        # 1. Header row
        ClauseMatrixExporter._write_header(ws)

        # 2. Tree data rows
        current_row = 2  # row 1 is the header
        merge_ranges: List[Dict] = []

        for idx, root_node in enumerate(document.structure, 1):
            current_row = ClauseMatrixExporter._write_node_recursive(
                ws,
                root_node,
                current_row,
                level_prefix=str(idx),
                merge_ranges=merge_ranges,
            )

        # 3. Merge cells for multi-clause sections
        ClauseMatrixExporter._merge_cells(ws, merge_ranges)

        # 4. Column widths
        ClauseMatrixExporter._set_column_widths(ws)

        # 5. Serialise
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def get_filename(doc_name: str) -> str:
        """
        Derive the Excel filename from the source document name.

        Args:
            doc_name: Original document file name.

        Returns:
            Filename such as ``"MyDoc_需求矩阵.xlsx"``.
        """
        if doc_name.lower().endswith(".pdf"):
            doc_name = doc_name[:-4]
        # Suffix is Chinese ("Requirements Matrix") – user-facing output.
        return f"{doc_name}_需求矩阵.xlsx"

    # ------------------------------------------------------------------
    # Header
    # ------------------------------------------------------------------

    @staticmethod
    def _write_header(ws) -> None:
        """Write the header row with styling."""
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid",
        )
        header_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        for col_idx, header in enumerate(ClauseMatrixExporter.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        ws.row_dimensions[1].height = 30

    # ------------------------------------------------------------------
    # Recursive row writer
    # ------------------------------------------------------------------

    @staticmethod
    def _write_node_recursive(
        ws,
        node: PageIndexNode,
        start_row: int,
        level_prefix: str,
        merge_ranges: List[Dict],
    ) -> int:
        """
        Recursively write node data into the worksheet.

        Args:
            ws: Active worksheet.
            node: Current tree node.
            start_row: First available row number.
            level_prefix: Hierarchy label (e.g. ``"1.2.3"``).
            merge_ranges: Accumulator for cell-merge descriptors.

        Returns:
            Next available row number.
        """
        current_row = start_row
        node_summary = node.summary or ""

        if node.clauses:
            first_row = current_row
            for clause in node.clauses:
                ClauseMatrixExporter._write_clause_row(
                    ws, current_row, level_prefix, node.title,
                    node_summary, clause,
                )
                current_row += 1

            # Merge the level / title / summary columns when multiple clauses
            if len(node.clauses) > 1:
                merge_ranges.append({
                    "start_row": first_row,
                    "end_row": current_row - 1,
                    "columns": [1, 2, 3],  # A, B, C
                })
        else:
            # Section-only row (no clauses)
            ClauseMatrixExporter._write_section_only_row(
                ws, current_row, level_prefix, node.title, node_summary,
            )
            current_row += 1

        # Recurse into children
        for idx, child in enumerate(node.nodes, 1):
            current_row = ClauseMatrixExporter._write_node_recursive(
                ws, child, current_row,
                level_prefix=f"{level_prefix}.{idx}",
                merge_ranges=merge_ranges,
            )

        return current_row

    # ------------------------------------------------------------------
    # Individual row writers
    # ------------------------------------------------------------------

    @staticmethod
    def _write_clause_row(
        ws,
        row: int,
        level: str,
        title: str,
        summary: str,
        clause: ClauseItem,
    ) -> None:
        """Write a data row that contains a clause."""
        border = _light_border()

        structured_info = " ".join(
            part for part in (clause.actor, clause.action, clause.object) if part
        )

        data = [
            level,
            title,
            summary,
            clause.matrix_id,
            clause.original_text[:100] if len(clause.original_text) > 100 else clause.original_text,
            clause.page_number,
            clause.type,
            structured_info,
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True,
            )
            # Center the hierarchy and page-number columns
            if col_idx in (1, 6):
                cell.alignment = Alignment(
                    horizontal="center", vertical="center",
                )

        ws.row_dimensions[row].height = 40

    @staticmethod
    def _write_section_only_row(
        ws,
        row: int,
        level: str,
        title: str,
        summary: str,
    ) -> None:
        """Write a section row without any clauses (grey background)."""
        border = _light_border()
        grey_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid",
        )

        data = [level, title, summary, "", "", "", "", ""]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True,
            )
            if col_idx <= 3:
                cell.fill = grey_fill
            if col_idx == 1:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center",
                )

        ws.row_dimensions[row].height = 30

    # ------------------------------------------------------------------
    # Cell merging & column widths
    # ------------------------------------------------------------------

    @staticmethod
    def _merge_cells(ws, merge_ranges: List[Dict]) -> None:
        """Merge cells for sections that span multiple clause rows."""
        for info in merge_ranges:
            for col_idx in info["columns"]:
                ws.merge_cells(
                    start_row=info["start_row"],
                    start_column=col_idx,
                    end_row=info["end_row"],
                    end_column=col_idx,
                )
                merged = ws.cell(row=info["start_row"], column=col_idx)
                h_align = "center" if col_idx == 1 else "left"
                merged.alignment = Alignment(
                    horizontal=h_align, vertical="center", wrap_text=True,
                )

    @staticmethod
    def _set_column_widths(ws) -> None:
        """Apply configured column widths."""
        for letter, width in ClauseMatrixExporter.COLUMN_WIDTHS.items():
            ws.column_dimensions[letter].width = width


# ---------------------------------------------------------------------------
# Module-private helpers
# ---------------------------------------------------------------------------

def _light_border() -> Border:
    """Return a thin light-grey border used for data rows."""
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


# Backward-compatible alias
ExcelExportService = ClauseMatrixExporter
