"""
Excel导出服务 - 将需求树导出为Excel文件
"""

from typing import List, Tuple, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from app.core.states import PageIndexDocument, PageIndexNode, RequirementItem


class ExcelExportService:
    """需求树Excel导出服务"""
    
    # 表头定义
    HEADERS = [
        "层级关系",
        "章节标题", 
        "章节概述",
        "需求ID",
        "需求",
        "页码",
        "风险提示",
        "建议应答方向"
    ]
    
    # 列宽设置（字符数）
    COLUMN_WIDTHS = {
        "A": 15,  # 层级关系
        "B": 25,  # 章节标题
        "C": 40,  # 章节概述
        "D": 20,  # 需求ID
        "E": 50,  # 需求
        "F": 10,  # 页码
        "G": 40,  # 风险提示
        "H": 40,  # 建议应答方向
    }
    
    @staticmethod
    def export_to_excel(document: PageIndexDocument) -> BytesIO:
        """
        将PageIndexDocument导出为Excel文件
        
        Args:
            document: PageIndex文档对象（包含需求树）
            
        Returns:
            BytesIO: Excel文件的字节流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "需求矩阵"
        
        # 1. 写入表头
        ExcelExportService._write_header(ws)
        
        # 2. 递归遍历树结构，写入数据
        current_row = 2  # 从第2行开始（第1行是表头）
        merge_ranges = []  # 记录需要合并的单元格范围
        
        for idx, root_node in enumerate(document.structure, 1):
            current_row = ExcelExportService._write_node_recursive(
                ws, 
                root_node, 
                current_row,
                level_prefix=str(idx),  # 顶层从1开始
                merge_ranges=merge_ranges  # 传递合并列表
            )
        
        # 3. 合并单元格
        ExcelExportService._merge_cells(ws, merge_ranges)
        
        # 4. 设置列宽
        ExcelExportService._set_column_widths(ws)
        
        # 5. 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def _write_header(ws):
        """写入表头并设置样式"""
        # 表头样式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_side = Side(style='thin', color="000000")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        for col_idx, header in enumerate(ExcelExportService.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 设置表头行高
        ws.row_dimensions[1].height = 30
    
    @staticmethod
    def _write_node_recursive(
        ws, 
        node: PageIndexNode, 
        start_row: int, 
        level_prefix: str,
        merge_ranges: List[Dict]
    ) -> int:
        """
        递归写入节点数据
        
        Args:
            ws: worksheet对象
            node: 当前节点
            start_row: 起始行号
            level_prefix: 层级前缀（如 "1", "1.1", "1.1.1"）
            merge_ranges: 记录需要合并的单元格范围
            
        Returns:
            int: 下一个可用的行号
        """
        current_row = start_row
        
        # 节点基础信息
        node_level = level_prefix
        node_title = node.title
        node_summary = node.summary or ""
        
        # 如果节点有需求，为每个需求创建一行
        if node.requirements:
            first_row = current_row  # 记录起始行
            
            for req in node.requirements:
                ExcelExportService._write_requirement_row(
                    ws, 
                    current_row,
                    node_level,
                    node_title,
                    node_summary,
                    req
                )
                current_row += 1
            
            # 如果有多个需求，记录前三列需要合并的范围
            if len(node.requirements) > 1:
                last_row = current_row - 1
                merge_ranges.append({
                    'start_row': first_row,
                    'end_row': last_row,
                    'columns': [1, 2, 3]  # 列A(1), B(2), C(3) - 层级、标题、概述
                })
        else:
            # 如果没有需求，只写入章节信息（需求相关列留空）
            ExcelExportService._write_section_only_row(
                ws,
                current_row,
                node_level,
                node_title,
                node_summary
            )
            current_row += 1
        
        # 递归处理子节点
        for idx, child_node in enumerate(node.nodes, 1):
            child_level_prefix = f"{level_prefix}.{idx}"
            current_row = ExcelExportService._write_node_recursive(
                ws,
                child_node,
                current_row,
                child_level_prefix,
                merge_ranges
            )
        
        return current_row
    
    @staticmethod
    def _write_requirement_row(
        ws,
        row: int,
        level: str,
        title: str,
        summary: str,
        req: RequirementItem
    ):
        """写入包含需求的行"""
        border_side = Side(style='thin', color="CCCCCC")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # 数据行
        data = [
            level,                          # 层级关系
            title,                          # 章节标题
            summary,                        # 章节概述
            req.matrix_id,                  # 需求ID
            req.requirement,                # 需求
            req.page_number,                # 页码
            req.risk_warning,               # 风险提示
            req.response_suggestion,        # 建议应答方向
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left", 
                vertical="top", 
                wrap_text=True
            )
            
            # 层级关系列居中
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 页码列居中
            if col_idx == 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置行高（自动调整）
        ws.row_dimensions[row].height = 40
    
    @staticmethod
    def _write_section_only_row(
        ws,
        row: int,
        level: str,
        title: str,
        summary: str
    ):
        """写入只有章节信息的行（无需求）"""
        border_side = Side(style='thin', color="CCCCCC")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # 浅灰色背景（区分无需求的章节）
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # 数据行（需求相关列留空）
        data = [
            level,      # 层级关系
            title,      # 章节标题
            summary,    # 章节概述
            "",         # 需求ID - 空
            "",         # 需求 - 空
            "",         # 页码 - 空
            "",         # 风险提示 - 空
            "",         # 建议应答方向 - 空
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left", 
                vertical="top", 
                wrap_text=True
            )
            
            # 前三列有背景色（表示这是章节行）
            if col_idx <= 3:
                cell.fill = gray_fill
            
            # 层级关系列居中
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[row].height = 30
    
    @staticmethod
    def _merge_cells(ws, merge_ranges: List[Dict]):
        """
        合并单元格
        
        Args:
            ws: worksheet对象
            merge_ranges: 需要合并的单元格范围列表
        """
        for merge_info in merge_ranges:
            start_row = merge_info['start_row']
            end_row = merge_info['end_row']
            columns = merge_info['columns']
            
            # 对每一列执行合并
            for col_idx in columns:
                # 合并单元格
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col_idx,
                    end_row=end_row,
                    end_column=col_idx
                )
                
                # 设置合并后单元格的对齐方式为垂直居中
                merged_cell = ws.cell(row=start_row, column=col_idx)
                
                # 保持原有的水平对齐，修改垂直对齐为居中
                if col_idx == 1:  # 层级关系列
                    merged_cell.alignment = Alignment(
                        horizontal="center", 
                        vertical="center", 
                        wrap_text=True
                    )
                else:  # 标题和概述列
                    merged_cell.alignment = Alignment(
                        horizontal="left", 
                        vertical="center", 
                        wrap_text=True
                    )
    
    @staticmethod
    def _set_column_widths(ws):
        """设置列宽"""
        for col_letter, width in ExcelExportService.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
    
    @staticmethod
    def get_filename(doc_name: str) -> str:
        """
        生成Excel文件名
        
        Args:
            doc_name: 文档名称
            
        Returns:
            str: 文件名（如 "招标文件_需求矩阵.xlsx"）
        """
        # 移除PDF后缀（不区分大小写）
        if doc_name.lower().endswith('.pdf'):
            doc_name = doc_name[:-4]
        
        return f"{doc_name}_需求矩阵.xlsx"

